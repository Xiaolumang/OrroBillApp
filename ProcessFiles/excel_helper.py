import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def expand_columns(output):
    wb = load_workbook(output)
    ws = wb.active
    ws.title = 'compare'
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set the width to slightly more than the maximum length
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def highlight_excel(output):
    # Load the Excel workbook and select a worksheet
    wb = load_workbook(output)
    ws = wb.active
    ws.title = 'summary'

    # Define a fill pattern for highlighting
    highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply highlight to each summary row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Assuming first row is header
        if row[4].value == 'Charge Back Journal':  # Check if it's a summary row
            for cell in row:
                cell.fill = highlight

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra space for better visibility
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output